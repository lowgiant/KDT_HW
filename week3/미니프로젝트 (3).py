'''
미니프로젝트 (3).ipynb 파일을 처음 보시는 분을 위해!

(1) 프로젝트 명 : 네이버 뉴스 자동 수집 후 메일 발송기

(2) 프로젝트 설명(200자 이내) : 특정 키워드에 해당하는 뉴스 기사를 네이버에서 찾아 기사 제목, 링크를 엑셀 파일로 만들어 미리 만들어둔 엑셀 파일에 있는 메일링 대상자에게 메일로 전달합니다.

(3) 프로젝트 과제 상세 :

사용자가 원하는 키워드 입력받기
네이버 뉴스를 수집해 주는 모듈을 이용해서 해당 키워드 뉴스 수집 후 엑셀 파일에 제목, 링크, 요약문 기록하기
수집 데이터 엑셀 파일을 메일링 대상자 엑셀 파일을 읽어 대상자들에게 메일 보내기
(4) 점검 및 합격 기준표 :

해당 파트를 수강하면서 사용한 문법과 기능을 활용할 수 있어야 합니다.
전체적인 업무의 흐름을 이해하고 작성하여야 합니다.
각 단계마다 주어진 과제 코드를 완성하여야 합니다.


아래 코드를 실행해서 NaverNewsCrawler 모듈을 임포트하세요.

# 크롤러 코드를 위한 모듈 설치
!pip install requests
!pip install beautifulsoup4
'''

#### 아래코드를 실행해 이메일 발송 기능에 필요한 모듈을 임포트하세요.
import email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re
import json
from openpyxl import Workbook, load_workbook
from NaverNewsCrawler import NaverNewsCrawler

keyword = input('수집할 기사 키워드 검색: ')
####사용자로 부터 기사 수집을 원하는 키워드를 input을 이용해 입력받아 ? 부분에 넣으세요
crawler = NaverNewsCrawler(keyword)

#### 수집한 데이터를 저장할 엑셀 파일명을 input을 이용해 입력받아 ? 부분에 넣으세요
excel_input_filename = input('저장할 엑셀 파일명(확장자X): ')
if excel_input_filename.count('.') > 1:
    excel_input_filename = ".".join(excel_input_filename.split('.')[0:excel_input_filename.count('.')])
    excel_filename = f'{excel_input_filename}.xlsx'
else:
    excel_filename = f'{excel_input_filename}.xlsx'

crawler.get_news(excel_filename)

#### gmail 발송 기능에 필요한 계정 정보를 아래 코드에 입력하세요.
with open('conf.json') as account_json:
    account_json = json.load(account_json)


SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USER = account_json['email']
SMTP_PASSWORD = account_json['password']

#### 아래 코드를 실행해 메일 발송에 필요한 send_mail 함수를 만드세요.
def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        # 엑셀 한글 파일명도 가능하도록 수정
        file_data.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()


#### 프로젝트 폴더에 있는 email_list.xlsx 파일에 이메일 받을 사람들의 정보를 입력하세요.
wb = Workbook()
ws = wb.active

people_count = 0

def email_list_people_add(name, email):
    global people_count 
    people_count += 1 
    ws[f'A{str(people_count)}'] = name
    ws[f'B{str(people_count)}'] = email

email_list_people_add('hsy', 'lowgiant@gmail.com')
email_list_people_add('hsy', 'newreview@naver.com')

wb.save('email_list.xlsx')

#### 엑셀 파일의 정보를 읽어올 수 있는 모듈을 import하세요.

#### email_list.xlsx 파일을 읽어와 해당 사람들에게 수집한 뉴스 정보 엑셀 파일을 send_mail 함수를 이용해 전송하세요.
wb = load_workbook('email_list.xlsx', read_only=True)
data = wb.active

subjects = f'{keyword}의 뉴스 입니다.'
contents = f'{keyword}에 대한 최근 뉴스 제목과 URL 내용을 첨부 파일로 보내드립니다.'

# 이메일 리스트에서 한명씩 메일 보냄
for row in data.iter_rows():    
    send_mail(row[0].value, row[1].value, subjects, contents, excel_filename)